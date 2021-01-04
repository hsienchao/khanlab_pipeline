#!/usr/bin/python
import re
import requests
import os,subprocess
import openpyxl
import sys
import argparse

def main(args):

    #User input:
    #sample ID list
    
    input_file = args.input.strip()
    sample_ids = args.sample.strip()
    library_col = args.library_col.strip()
    run_col = args.run_col.strip()
    samples = sample_ids.split(",")
    libs = {}
    runs = {}
    prefix = "Sample_"
    print("#Input")
    for sample_id in samples:
        # remove Sample_ if any
        sample_id = sample_id[sample_id.startswith(prefix) and len(prefix):]
        idx = sample_id.rfind('_')
        if idx == -1:
            sys.exit("invalid sample_id:" + sample_id)
        library_id = sample_id[0:idx].strip()
        run_id = sample_id[idx+1:].strip()
        if run_id == "":
            sys.exit("invalid sample_id:" + sample_id )
        print(library_id + "\t" + run_id)
        libs[library_id] = run_id
    xfile = openpyxl.load_workbook(input_file, data_only=True)
    ws= xfile.worksheets[0]
    lib_id_idx = 0;
    run_id_idx = 0;
    for i in range(ws.max_column-1):
        if ws.cell(row=1, column=i+1).value != None:
            col_value = ws.cell(row=1, column=i+1).value.strip()
            if col_value.strip() == library_col:
                lib_id_idx = i + 1;
            if col_value.strip() == run_col:
                run_id_idx = i + 1;
    print("#Updated")
    for i in range(1,ws.max_row-1):
        ws_lib_id = ws.cell(row=i+1, column=lib_id_idx).value
        #print(ws_lib_id)
        if ws_lib_id != None:
            run_id = libs.get(ws_lib_id.strip())
            if run_id != None:
                ws_run_id = ws.cell(row=i+1, column=run_id_idx).value
                if ws_run_id != None:
                    if ws_run_id.strip() == "TBD":
                        print(ws_lib_id + "\t" + ws_run_id + "\t" + run_id)
                        #print("replace TBD of " + ws_lib_id + " with " + run_id)
                        runs[str(i+1)] = run_id
    #print(lib_id_idx)
    #print(run_id_idx)
    xfile.close()
    xfile = openpyxl.load_workbook(args.input.strip())
    ws= xfile.worksheets[0]
    for i in runs:
        print("saving " + str(i))
        ws.cell(row=int(i), column=run_id_idx).value = runs[i]
    #for i in range(ws.max_row):
    #    ws.cell(row=i+1, column=run_id_idx).number_format = '@'
    #print(i + " " + runs[i])
        
    xfile.save(input_file)
    xfile.close()

parser = argparse.ArgumentParser(description='Replace Flow cell/SRR ID.')
parser.add_argument("--sample", "-s", metavar="SAMPLE_ID", required=True, help="Required. SAMPLE_ID ID list (comma seperated).")
parser.add_argument("--out_format", "-f", metavar="OUTPUT", help="Output format: table or yaml. default: %(default)s", default="table")
parser.add_argument("--input", "-i", metavar="Input master Excel file", help="Master Excel file. default: %(default)s", default="/data/khanlab/projects/ChIP_seq/manage_samples/ChIP_seq_samples.xlsx")
parser.add_argument("--library_col", "-l", metavar="Library ID column name", help="Library ID column name. default: %(default)s", default="Amplified_Sample_Library_Name")
parser.add_argument("--run_col", "-r", metavar="Run ID column name", help="Run ID column name. default: %(default)s", default="SequencingRun_GEO")
args = parser.parse_args()

main(args)